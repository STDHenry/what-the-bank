from flask import Flask, request, redirect, url_for
import os
import random
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

EXCEL_FILE = "bank_database.xlsx"
CURRENT_OTP = ""
PENDING_USER_DATA = {}

def init_excel_file():
    """Tạo file Excel cơ sở dữ liệu nếu chưa tồn tại trên máy"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        # Thiết lập 5 cột lưu trữ thông tin chuẩn
        ws.append(["Tên", "Mật khẩu", "Ngày sinh", "Số điện thoại", "Số dư"])
        wb.save(EXCEL_FILE)

# Ép hệ thống tạo sẵn file Excel ngay khi vừa bật lệnh chạy
init_excel_file()

def register_user_excel(user, pwd, dob, phone, balance=500000):
    """Ghi thông tin tài khoản mới đăng ký kèm 500k VND vào file Excel"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    
    # Quét qua bảng tính Excel để kiểm tra trùng lặp Tên hoặc SĐT
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 4:
            if row[0] == user or row[3] == phone:
                return False
                
    # Thêm một hàng dữ liệu mới vào bảng tính Excel
    ws.append([user, pwd, dob, phone, balance])
    wb.save(EXCEL_FILE)
    return True
def verify_login_excel(login_id, pwd):
    """Kiểm tra tài khoản (Tên hoặc SĐT) và Mật khẩu trong Excel"""
    if not os.path.exists(EXCEL_FILE):
        return None
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 4:
            # So khớp nếu (Cột 0 trùng ID hoặc Cột 3 trùng ID) và Cột 1 trùng Mật khẩu
            if (str(row[0]) == login_id or str(row[3]) == login_id) and str(row[1]) == pwd:
                return row[0]  # Trả về Tên tài khoản chuẩn
    return None

def get_balance_excel(user):
    """Lấy số dư từ cột số 5 (vị trí index 4 trong mảng row)"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 5 and row[0] == user:
            return int(row[4])
    return 0

def update_balance_excel(user, new_balance):
    """Cập nhật số dư mới sau khi chơi chứng khoán IA FEX"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == user:
            row[4].value = new_balance  # Ghi số tiền mới vào cột 5
            wb.save(EXCEL_FILE)
            return True
    return False

def process_transfer_excel(sender, receiver, amount):
    """Trừ tiền người gửi và cộng tiền người nhận trực tiếp trong file Excel"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Users"]
    
    sender_row = None
    receiver_row = None
    
    # Tìm kiếm 2 hàng dữ liệu của người gửi và người nhận
    for row in ws.iter_rows(min_row=2):
        if row[0].value == sender:
            sender_row = row
        if row[0].value == receiver:
            receiver_row = row
            
    if not receiver_row:
        return "Lỗi: Người nhận không tồn tại trên hệ thống!"
    if sender == receiver:
        return "Lỗi: Không thể tự chuyển tiền cho chính mình!"
        
    sender_balance = int(sender_row[4].value)
    if sender_balance < amount:
        return "Lỗi: Số dư tài khoản không đủ!"
        
    # Tính toán số dư mới và lưu lại vào các ô của hàng Excel tương ứng
    sender_row[4].value = sender_balance - amount
    receiver_row[4].value = int(receiver_row[4].value) + amount
    wb.save(EXCEL_FILE)
    return "Thành công"

def log_otp_excel(user, otp):
    """Ghi nhật ký mã OTP bảo mật vào một Sheet riêng để theo dõi"""
    wb = load_workbook(EXCEL_FILE)
    if "OTP_Logs" not in wb.sheetnames:
        wb.create_sheet("OTP_Logs")
    ws = wb["OTP_Logs"]
    ws.append([user, f"Mã OTP hệ thống cấp: {otp}"])
    wb.save(EXCEL_FILE)
# ==========================================
# CHUNG: ĐOẠN MÃ STYLE CSS CHO TRANG AUTH (XÁM TỐI + XANH HẢI QUÂN)
# ==========================================
AUTH_STYLE = """
<style>
    body {
        margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif;
        background-color: #202124; /* Nền xám tối chủ đạo */
        color: #ffffff;
        display: flex; justify-content: center; align-items: center;
        height: 100vh;
    }
    .auth-card {
        background-color: #000080; /* Màu xanh hải quân hải quân #000080 */
        padding: 40px; border-radius: 12px;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.4);
        width: 360px; text-align: center;
    }
    .auth-card h1 { margin: 0 0 5px 0; font-size: 30px; letter-spacing: 1px; }
    .auth-card p.subtitle { color: #bdc3c7; font-size: 13px; margin-bottom: 30px; }
    .input-field {
        width: 93%; padding: 11px; margin-bottom: 15px;
        border: 1px solid #bdc3c7; border-radius: 6px;
        font-size: 15px; background: #ffffff; color: #333333;
    }
    .btn-link {
        display: block; width: 100%; padding: 12px 0; border-radius: 6px;
        font-size: 16px; font-weight: bold; text-decoration: none;
        transition: all 0.2s ease-in-out; margin-bottom: 12px;
    }
    /* Hiệu ứng di chuột: phóng to lên 1 chút và tối màu đi */
    .btn-link:hover {
        transform: scale(1.05);
        filter: brightness(0.85);
    }
    .btn-submit {
        width: 100%; padding: 12px; border: none; border-radius: 6px;
        font-size: 16px; font-weight: bold; cursor: pointer;
        transition: all 0.2s ease-in-out;
    }
    .btn-submit:hover {
        transform: scale(1.05);
        filter: brightness(0.85);
    }
    .error-box { color: #ff4757; background: rgba(255, 71, 87, 0.1); padding: 10px; border-radius: 6px; margin-bottom: 15px; font-size: 14px; font-weight: bold; }
</style>
"""

# ==========================================
# 1. GIAO DIỆN CHỌN ĐĂNG NHẬP / ĐĂNG KÝ BAN ĐẦU
# ==========================================
@app.route("/")
def index():
    msg = request.args.get("msg", "")
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>What The Bank</title>
        {AUTH_STYLE}
    </head>
    <body>
        <div class="auth-card">
            <h1>What The Bank</h1>
            <p class="subtitle">Hệ thống mô phỏng tài chính tối tân</p>
            {f'<div class="error-box">⚠️ {msg}</div>' if msg else ''}
            <a href="/login" class="btn-link" style="background:#1289A7; color:white;">ĐĂNG NHẬP</a>
            <a href="/register" class="btn-link" style="background:#2ed573; color:white;">ĐĂNG KÝ TÀI KHOẢN</a>
        </div>
    </body>
    </html>
    """

# ==========================================
# 2. GIAO DIỆN FORM ĐĂNG KÝ ĐẦY ĐỦ THÔNG TIN
# ==========================================
@app.route("/register")
def register():
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>WTB - Đăng ký</title>
        {AUTH_STYLE}
    </head>
    <body>
        <div class="auth-card" style="text-align: left;">
            <h2 style="margin-top:0; text-align:center;">Đăng Ký Thành Viên</h2>
            <form action="/auth-submit?action=reg" method="POST">
                <input type="text" name="username" class="input-field" placeholder="Tên tài khoản" required>
                <input type="password" name="password" class="input-field" placeholder="Mật khẩu bảo mật" required>
                <label style="font-size: 13px; color: #bdc3c7; display:block; margin-bottom:4px;">Ngày tháng năm sinh:</label>
                <input type="date" name="dob" class="input-field" required>
                <input type="tel" name="phone" class="input-field" placeholder="Số điện thoại" required>
                <button type="submit" class="btn-submit" style="background-color: #2ed573; color: white;">TIẾP TỤC NHẬN MÃ OTP</button>
            </form>
            <a href="/" style="color:#bdc3c7; text-decoration:none; font-size:13px; display:block; text-align:center; margin-top:15px;">Quay lại</a>
        </div>
    </body>
    </html>
    """

# ==========================================
# 3. GIAO DIỆN FORM ĐĂNG NHẬP (TÊN HOẶC SĐT)
# ==========================================
@app.route("/login")
def login():
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>WTB - Đăng nhập</title>
        {AUTH_STYLE}
    </head>
    <body>
        <div class="auth-card" style="text-align: left;">
            <h2 style="margin-top:0; text-align:center;">Đăng Nhập Hệ Thống</h2>
            <form action="/auth-submit?action=login" method="POST">
                <input type="text" name="login_id" class="input-field" placeholder="Tên tài khoản hoặc Số điện thoại" required>
                <input type="password" name="password" class="input-field" placeholder="Mật khẩu" required>
                <button type="submit" class="btn-submit" style="background-color: #1289A7; color: white;">TIẾP TỤC NHẬN MÃ OTP</button>
            </form>
            <a href="/" style="color:#bdc3c7; text-decoration:none; font-size:13px; display:block; text-align:center; margin-top:15px;">Quay lại</a>
        </div>
    </body>
    </html>
    """
# ==========================================
# 4. GIAO DIỆN BẢNG MÃ OTP XÁC THỰC BẢO MẬT
# ==========================================
@app.route("/otp-verify")
def otp_verify():
    global CURRENT_OTP
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>WTB - Xác thực OTP</title>
        {AUTH_STYLE}
    </head>
    <body>
        <div class="auth-card">
            <h2>Xác Thực OTP</h2>
            <div style="background: #ffffff; padding: 12px; font-size: 24px; font-weight: bold; color: #dc3545; margin: 15px 0; border-radius: 6px; letter-spacing: 2px;">
                MÃ FILE: {CURRENT_OTP}
            </div>
            <form action="/otp-check" method="POST">
                <input type="text" name="otp_input" class="input-field" placeholder="Nhập lại mã OTP ở trên" required style="text-align:center;">
                <button type="submit" class="btn-submit" style="background-color: #2ed573; color: white;">XÁC NHẬN VÀO HỆ THỐNG</button>
            </form>
        </div>
    </body>
    </html>
    """

# ==========================================
# 5. BIÊN DỊCH DỮ LIỆU ĐỂ QUYẾT ĐỊNH CẤP MÃ OTP
# ==========================================
@app.route("/auth-submit", methods=["POST"])
def auth_submit():
    global CURRENT_OTP, PENDING_USER_DATA
    action = request.args.get("action")
    
    CURRENT_OTP = str(random.randint(100000, 999999))
    
    if action == "reg":
        user = request.form.get("username").strip()
        pwd = request.form.get("password").strip()
        dob = request.form.get("dob")
        phone = request.form.get("phone").strip()
        
        # Đọc file Excel kiểm tra trùng lặp thông tin trước khi cấp OTP
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Users"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:
                    if row[0] == user or row[3] == phone:
                        return redirect(url_for("index", msg="Lỗi: Tên hoặc SĐT đã tồn tại!"))
                        
        PENDING_USER_DATA = {"action": "reg", "user": user, "pwd": pwd, "dob": dob, "phone": phone, "balance": 500000}
        
    elif action == "login":
        login_id = request.form.get("login_id").strip()
        pwd = request.form.get("password").strip()
        
        found_user = verify_login_excel(login_id, pwd)
        if not found_user:
            return redirect(url_for("index", msg="Lỗi: Sai thông tin đăng nhập!"))
            
        PENDING_USER_DATA = {"action": "login", "user": found_user}

    # Ghi lại dấu vết mã OTP vừa tạo vào trang nhật ký riêng trong Excel
    log_otp_excel(PENDING_USER_DATA["user"], CURRENT_OTP)
    return redirect(url_for("otp_verify"))

# ==========================================
# 6. KIỂM TRA MÃ OTP NGƯỜI DÙNG NHẬP VÀO
# ==========================================
@app.route("/otp-check", methods=["POST"])
def otp_check():
    global CURRENT_OTP, PENDING_USER_DATA
    otp_input = request.form.get("otp_input").strip()
    
    if otp_input != CURRENT_OTP:
        return redirect(url_for("index", msg="Xác thực OTP thất bại! Vui lòng thử lại."))
        
    if PENDING_USER_DATA["action"] == "reg":
        register_user_excel(
            PENDING_USER_DATA["user"], 
            PENDING_USER_DATA["pwd"], 
            PENDING_USER_DATA["dob"], 
            PENDING_USER_DATA["phone"], 
            PENDING_USER_DATA["balance"]
        )
        
    return redirect(url_for("dashboard", username=PENDING_USER_DATA["user"]))
# ==========================================
# 7. GIAO DIỆN CHÍNH (DASHBOARD NGÂN HÀNG - XÁM TỐI & XANH HẢI QUÂN)
# ==========================================
@app.route("/dashboard")
def dashboard():
    user = request.args.get("username")
    msg = request.args.get("msg", "")
    balance = get_balance_excel(user)
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>What The Bank - Dashboard</title>
        <style>
            body {{
                margin: 0; padding: 0; font-family: 'Segoe UI', Arial, sans-serif;
                background-color: #202124; color: #ffffff;
            }}
            /* Thanh điều hướng màu xanh hải quân chuẩn #000080 */
            .navbar {{
                background-color: #000080; color: white; padding: 15px 30px;
                display: flex; justify-content: space-between; align-items: center;
                box-shadow: 0 4px 12px rgba(0,0,0,0.3);
            }}
            .navbar h2 {{ margin: 0; font-size: 24px; letter-spacing: 1px; }}
            
            /* Khu vực chứa Avatar nhấp chuột thả bảng cài đặt */
            .avatar-container {{ position: relative; display: inline-block; cursor: pointer; }}
            .avatar {{
                width: 45px; height: 45px; border-radius: 50%; border: 2px solid white;
                background-image: url('/static/logo.png'); background-size: cover; background-position: center;
                transition: all 0.2s ease;
            }}
            .avatar:hover {{ transform: scale(1.05); filter: brightness(0.85); }}
            
            /* Bảng lựa chọn Đăng xuất thả xuống (Dropdown) */
            .dropdown {{
                display: none; position: absolute; right: 0; top: 55px;
                background: #ffffff; box-shadow: 0 4px 15px rgba(0,0,0,0.4);
                border-radius: 6px; width: 120px; text-align: center; z-index: 10;
                overflow: hidden;
            }}
            .dropdown a {{
                display: block; padding: 12px; color: #dc3545; text-decoration: none;
                font-size: 14px; font-weight: bold; transition: background 0.2s;
            }}
            .dropdown a:hover {{ background: #f1f2f6; }}
            
            .container {{ max-width: 600px; margin: 30px auto; padding: 0 20px; text-align: center; }}
            .card {{
                background-color: #2d2d30; border-radius: 12px; padding: 25px;
                box-shadow: 0 4px 15px rgba(0,0,0,0.2); border: 1px solid #3e3e42;
                margin-bottom: 20px; text-align: left;
            }}
            .balance {{ font-size: 32px; font-weight: bold; color: #2ed573; margin: 15px 0; text-align: center; }}
            
            /* Cụm nút bấm chuyển đổi tab dịch vụ mới dạng Icon theo yêu cầu */
            .service-tabs {{
                display: flex; justify-content: center; gap: 20px; margin: 25px 0;
            }}
            .icon-btn {{
                padding: 12px 25px; font-size: 18px; font-weight: bold; color: white;
                border: none; border-radius: 8px; cursor: pointer;
                transition: all 0.2s ease-in-out; display: flex; align-items: center; gap: 8px;
            }}
            /* Hiệu ứng di chuột chạm vào to lên và tối đi */
            .icon-btn:hover {{ transform: scale(1.05); filter: brightness(0.85); }}
            .btn-transfer-tab {{ background-color: #1289A7; }}
            .btn-fex-tab {{ background-color: #f1c40f; color: black; }}
            
            .input-box {{
                width: 95%; padding: 11px; margin-bottom: 12px;
                border: 1px solid #555; border-radius: 6px;
                font-size: 14px; background: #3c3c3c; color: white;
            }}
            .input-box::placeholder {{ color: #aaa; }}
            
            .action-btn {{
                width: 100%; padding: 12px; background: #1289A7; color: white;
                border: none; border-radius: 6px; font-size: 15px; font-weight: bold;
                cursor: pointer; transition: all 0.2s;
            }}
            .action-btn:hover {{ transform: scale(1.03); filter: brightness(0.85); }}
            .status-msg {{ color: #f1c40f; font-weight: bold; margin-bottom: 15px; text-align: center; }}
        </style>
        <script>
            function toggleDropdown() {{
                var menu = document.getElementById("avatarMenu");
                menu.style.display = (menu.style.display === "block") ? "none" : "block";
            }}
            
            // Hàm JavaScript giúp bấm nút Icon đổi ngay giao diện dịch vụ mà không cần load lại trang
            function switchService(tabName) {{
                document.getElementById('transfer-box').style.display = tabName === 'transfer' ? 'block' : 'none';
                document.getElementById('fex-box').style.display = tabName === 'fex' ? 'block' : 'none';
            }}
        </script>
    </head>
    <body>

        <div class="navbar">
            <h2>What The Bank</h2>
            <div class="avatar-container" onclick="toggleDropdown()">
                <div class="avatar" title="Nhấp để Đăng xuất"></div>
                <div id="avatarMenu" class="dropdown">
                    <a href="/">Đăng xuất</a>
                </div>
            </div>
        </div>

        <div class="container">
            <h2 style="text-align: left; color: #bdc3c7;">Tài khoản: <span style="color: #ffffff;">{user}</span></h2>
            {f'<p class="status-msg">ℹ️ {msg}</p>' if msg else ''}
            
            <!-- THẺ THÔNG TIN SỐ DƯ -->
            <div class="card" style="text-align: center;">
                <h3 style="margin-top: 0; color: #bdc3c7;">Số dư hiện tại</h3>
                <div class="balance">{balance:,} VND</div>
            </div>

            <!-- CỤM NÚT ICON CHỌN DỊCH VỤ -->
            <div class="service-tabs">
                <button class="icon-btn btn-transfer-tab" onclick="switchService('transfer')">$ ↓ Chuyển Tiền</button>
                <button class="icon-btn btn-fex-tab" onclick="switchService('fex')">↑ $ Chứng Khoán</button>
            </div>

            <!-- KHỐI DỊCH VỤ 1: CHUYỂN TIỀN NỘI BỘ -->
            <div id="transfer-box" class="card" style="display: block;">
                <h3 style="margin-top: 0; color: #1289A7; border-bottom: 1px solid #3e3e42; padding-bottom: 10px;">Chuyển Tiền Nội Bộ</h3>
                <form action="/transfer-submit" method="POST" style="margin-top: 15px;">
                    <input type="hidden" name="sender" value="{user}">
                    <input type="text" name="receiver" class="input-box" placeholder="Tên tài khoản người nhận" required><br>
                    <input type="number" name="amount" class="input-box" placeholder="Số tiền cần chuyển (VND)" required><br>
                    <button type="submit" class="action-btn">XÁC NHẬN CHUYỂN KHOẢN</button>
                </form>
            </div>

            <!-- VỊ TRÍ ĐỢT 6 SẼ CHÈN KHỐI GIAO DIỆN SÀN CHỨNG KHOÁN IA FEX VÀO ĐÂY -->

        </div>
    </body>
    </html>
    """

# ==========================================
# 8. ĐƯỜNG DẪN XỬ LÝ CHUYỂN KHOẢN TRONG FILE EXCEL
# ==========================================
@app.route("/dashboard-fex-addon")
def dashboard_fex_addon():
    """Hàm nâng cấp đồ thị hình nến (Candlestick) và các khối tương tác thời gian thực cho sàn IA FEX"""
    user = request.args.get("username")
    msg = request.args.get("msg", "")
    balance = get_balance_excel(user)
    
    fex_html_card = f"""
            <div id="fex-box" class="card" style="display: none;">
                <h3 style="margin-top: 0; color: #f1c40f; border-bottom: 1px solid #3e3e42; padding-bottom: 10px; display: flex; justify-content: space-between; align-items: center;">
                    <span>Sàn Chứng Khoán Hình Khối: IA FEX 📊</span>
                    <span id="ticker-price" style="color: #2ed573; font-family: monospace; font-size: 22px;">120.00 USD</span>
                </h3>
                <div style="display: flex; justify-content: space-between; font-size: 12px; color: #aaa; margin-top: -5px; margin-bottom: 10px;">
                    <span>Xu hướng thị trường: <b id="market-trend">ỔN ĐỊNH</b></span>
                    <span>Khối lượng GD: <b id="market-volume">0k</b></span>
                </div>
                
                <!-- Khung vẽ đồ thị hình khối nến -->
                <canvas id="fexChart" width="540" height="220" style="width: 100%; background: #15161e; border-radius: 8px; border: 1px solid #383a40; margin-bottom: 15px;"></canvas>
                
                <!-- Form Đặt Lệnh Tương Tác Tài Chính -->
                <form action="/fex-trade" method="POST" style="display: flex; gap: 10px; align-items: center;">
                    <input type="hidden" name="username" value="{user}">
                    <input type="number" name="trade_amount" class="input-box" placeholder="Số tiền đầu tư (VND)" required style="margin-bottom:0; flex:1; border-color: #f1c40f;">
                    <button type="submit" name="direction" value="up" class="action-btn" style="background: #2ed573; width: 110px; box-shadow: 0 4px 10px rgba(46,213,115,0.3);">MUA $ ↑</button>
                    <button type="submit" name="direction" value="down" class="action-btn" style="background: #ff4757; width: 110px; box-shadow: 0 4px 10px rgba(255,71,87,0.3);">BÁN $ ↓</button>
                </form>
            </div>

            <script>
                const canvas = document.getElementById('fexChart');
                const ctx = canvas.getContext('2d');
                const priceTicker = document.getElementById('ticker-price');
                const trendTicker = document.getElementById('market-trend');
                const volumeTicker = document.getElementById('market-volume');
                
                let candles = []; // Lưu danh sách các khối nến
                let currentPrice = 120;

                // Khởi tạo sẵn 15 cây nến ban đầu cho đồ thị có dữ liệu ngay
                for(let i=0; i<15; i++) {{
                    let openPrice = currentPrice;
                    let closePrice = currentPrice + (Math.random() * 30 - 15);
                    let highPrice = Math.max(openPrice, closePrice) + (Math.random() * 10);
                    let lowPrice = Math.min(openPrice, closePrice) - (Math.random() * 10);
                    candles.push({{ open: openPrice, close: closePrice, high: highPrice, low: lowPrice, vol: Math.floor(Math.random()*90 + 10) }});
                    currentPrice = closePrice;
                }}

                function drawCandleChart() {{
                    // 1. Tạo biến động giá ngẫu nhiên kịch tính từng giây
                    let openPrice = currentPrice;
                    let change = (Math.random() * 40 - 20); // Giật từ -20 đến +20 USD
                    currentPrice = Math.max(10, currentPrice + change);
                    let closePrice = currentPrice;
                    
                    // Tạo râu nến (giá cao nhất / thấp nhất trong giây đó)
                    let highPrice = Math.max(openPrice, closePrice) + (Math.random() * 8);
                    let lowPrice = Math.max(5, Math.min(openPrice, closePrice) - (Math.random() * 8));
                    let randomVol = Math.floor(Math.random() * 85 + 15);

                    // Đẩy khối nến mới vào mảng dữ liệu
                    candles.push({{ open: openPrice, close: closePrice, high: highPrice, low: lowPrice, vol: randomVol }});
                    if(candles.length > 18) candles.shift(); // Giữ tối đa 18 khối nến trên màn hình

                    // 2. Cập nhật bảng điện tử tương tác
                    priceTicker.innerText = currentPrice.toFixed(2) + " USD";
                    volumeTicker.innerText = randomVol + "k USD";
                    if(change >= 0) {{
                        priceTicker.style.color = "#2ed573";
                        trendTicker.innerText = "TĂNG TRƯỞNG 🟢";
                        trendTicker.style.color = "#2ed573";
                    }} else {{
                        priceTicker.style.color = "#ff4757";
                        trendTicker.innerText = "SUY THOÁI 🔴";
                        trendTicker.style.color = "#ff4757";
                    }}

                    // 3. Xóa khung cũ và vẽ lại toàn bộ đồ thị hình khối nến
                    ctx.clearRect(0, 0, canvas.width, canvas.height);
                    
                    // Vẽ các đường lưới ngang cố định
                    ctx.strokeStyle = '#222531';
                    ctx.lineWidth = 1;
                    for(let i = 30; i < canvas.height - 30; i += 40) {{
                        ctx.beginPath(); ctx.moveTo(0, i); ctx.lineTo(canvas.width, i); ctx.stroke();
                    }}

                    // Tính toán khoảng cách căn lề giữa các khối nến
                    let padding = 15;
                    let candleWidth = (canvas.width - (padding * 2)) / candles.length - 8;
                    let spacing = (canvas.width - (padding * 2)) / candles.length;

                    // Vòng lặp vẽ từng khối nến một
                    for(let i = 0; i < candles.length; i++) {{
                        let c = candles[i];
                        let isGreen = c.close >= c.open;
                        let color = isGreen ? '#2ed573' : '#ff4757';
                        
                        // Hàm chuyển đổi giá trị số sang tọa độ Y trên màn hình vẽ
                        function getBoxY(val) {{
                            return canvas.height - 40 - ((val / 300) * (canvas.height - 80));
                        }}

                        let yOpen = getBoxY(c.open);
                        let yClose = getBoxY(c.close);
                        let yHigh = getBoxY(c.high);
                        let yLow = getBoxY(c.low);
                        
                        let x = padding + (i * spacing);
                        let xCenter = x + (candleWidth / 2);

                        // Vẽ râu nến (đường thẳng từ đỉnh cao nhất đến đáy thấp nhất)
                        ctx.strokeStyle = color;
                        ctx.lineWidth = 2;
                        ctx.beginPath();
                        ctx.moveTo(xCenter, yHigh);
                        ctx.lineTo(xCenter, yLow);
                        ctx.stroke();

                        // Vẽ Thân nến (Hình khối chữ nhật đặc)
                        ctx.fillStyle = color;
                        let bodyHeight = Math.abs(yClose - yOpen);
                        let yTop = Math.min(yOpen, yClose);
                        // Nếu thân nến quá mỏng, ép hiển thị 2px cho rõ khối
                        if(bodyHeight < 2) bodyHeight = 2; 
                        ctx.fillRect(x, yTop, candleWidth, bodyHeight);

                        // Vẽ thêm Khối lượng giao dịch (Volume) dạng cột mờ ở đáy đồ thị
                        ctx.fillStyle = isGreen ? 'rgba(46, 213, 115, 0.15)' : 'rgba(255, 71, 87, 0.15)';
                        let volHeight = (c.vol / 100) * 35; // Chiều cao cột volume tối đa 35px
                        ctx.fillRect(x, canvas.height - volHeight, candleWidth, volHeight);
                    }}
                }}

                // Bật chế độ tự động giật khối đồ thị 1 giây một lần
                setInterval(drawCandleChart, 1000);
                drawCandleChart();
            </script>
    """
    return app.view_functions['dashboard']().replace("<!-- VỊ TRÍ ĐỢT 5 SẼ CHÈN THÊM THẺ CHỨNG KHOÁN IA FEX VÀO ĐÂY -->", fex_html_card)

@app.route("/dashboard")
def dashboard_override():
    return app.view_functions['dashboard_fex_addon']()

@app.route("/fex-trade", methods=["POST"])
def fex_trade():
    user = request.form.get("username")
    trade_amount = int(request.form.get("trade_amount"))
    direction = request.form.get("direction")
    
    balance = get_balance_excel(user)
    if balance < trade_amount:
        return redirect(url_for("dashboard_override", username=user, msg="Lỗi FEX: Số dư tài khoản không đủ để đặt lệnh!"))
        
    is_win = random.choice([True, False])
    
    if is_win:
        new_balance = balance + trade_amount  
        result_msg = f"Sàn IA FEX: Khớp lệnh thành công! Bạn đã thắng lệnh đầu tư và kiếm được +{trade_amount:,} VND."
    else:
        new_balance = balance - trade_amount  
        result_msg = f"Sàn IA FEX: Thị trường quét râu nến đảo chiều! Bạn đã thua lệnh đầu tư và bị trừ -{trade_amount:,} VND."
        
    update_balance_excel(user, new_balance)
    return redirect(url_for("dashboard_override", username=user, msg=result_msg))

if __name__ == "__main__":
    # Đổi sang host 0.0.0.0 để mở cổng ra mạng Internet toàn cầu
    app.run(host='0.0.0.0', port=5000)

